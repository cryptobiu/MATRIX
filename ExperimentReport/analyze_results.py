import sys
import os
import glob
import json
import time
import smtplib
from os.path import expanduser, exists
from os.path import basename
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


config_file_path = sys.argv[1]
task_idx = sys.argv[2]

results_path = input('Enter results directory. current path is: %s): ' % os.getcwd())

with open(config_file_path) as conf_file:
    conf_data = json.load(conf_file, object_pairs_hook=OrderedDict)
    remote_directory = conf_data['workingDirectory']


protocol_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
protocol_name = conf_data['protocol']

results_file_name = 'ExperimentReport/Results_%s_%s.xlsx' % (protocol_name, protocol_time)
style1 = NamedStyle(number_format='#.##')
# style1 = wb.add_format({'num_format': '#.##'})


def send_email():

    users = list(conf_data['emails'].values())
    configurations = list(conf_data['configurations'].values())
    regions = list(conf_data['regions'].values())
    address_me = 'biu.cyber.experiments@gmail.com'
    me = 'BIU Cyber Experiments <biu.cyber.experiments@gmail.com>'

    message = MIMEMultipart()
    message['Subject'] = 'Experiment results for protocol %s' % protocol_name
    message['From'] = me
    message['To'] = ', '.join(users)
    message_body = 'Results for protocol %s are attached.\n' % protocol_name
    message_body += 'The configuration(s) for this experiment are:\n\n'

    # write all the configuration to mail
    for conf in configurations:
        vals = conf.split('@')
        values_str = ''

        for val in vals:
            values_str += '%s ' % val
        message_body += '%s\n' % values_str
    # write all regions to mail
    message_body += 'The region(s) the experiment executed are:\n\n'
    for region in regions:
        message_body += '%s\n' % region

    message_body += '\nBIU Cyber Experiments'
    message.attach(MIMEText(message_body))

    # attach to mail all the reports file
    results_files = glob.glob('ExperimentReport/Results_%s_%s_*.xlsx' % (protocol_name, protocol_time))
    for file in results_files:
        with open(file, 'rb') as fli:
            part = MIMEApplication(fli.read(), Name=basename(file))
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(file)
            message.attach(part)

    server = smtplib.SMTP('smtp.gmail.com:587')
    server.starttls()
    server.login(address_me, 'Cyberexp1!')
    server.sendmail(me, users, message.as_string())
    server.quit()


def upload_to_git():
    working_dir = '%s/ExperimentsResults/%s' % (expanduser('~'), protocol_name)
    # create working directory for the protocol if not exist

    os.system('mkdir -p %s ' % working_dir)

    # move the results folder and the xlsx file to the protocol folder

    os.system('mv %s %s' % (results_path, working_dir))
    os.system('mv %s %s' % (results_file_name, working_dir))

    # Upload data to git

    os.curdir(expanduser('~/ExperimentsResults'))
    os.system('git add %s/%s %s/%s' % (protocol_name, results_path, protocol_name, results_file_name))
    os.system('git commit -m "Add results for protocol %s' % protocol_name)
    os.system('git push git@github.com:cryptobiu/ExperimentsResults.git')


def analyze_results(files_list, analysis_type):

    parties = set()
    for file in files_list:
        parties.add(int(basename(file.split('_')[3].split('.')[0].split('=')[1])))

    parties = list(parties)
    parties.sort()

    # parsing tasks names from json file
    # Assumption : all the parties measure the same tasks

    tasks_names = dict()

    # load one of the data files to receive the headers to the xlsx file

    with open(files_list[0], 'r') as f:
        data = json.load(f)

    # init list values
    for i in range(len(data)):
        tasks_names[data[i]['name']] = list()

    num_of_repetitions = conf_data['numOfInternalRepetitions']

    if not exists(results_file_name):
        wb = Workbook(write_only=False)
        ws = wb.active
        wb.remove(ws)
    else:
        wb = load_workbook(results_file_name)

    ws = wb.create_sheet(analysis_type)
    ws.cell(row=1, column=1, value='Phase/Number of Parties')

    files_list.sort()

    # counter = 0
    for party_idx in range(len(parties)):
        ws.cell(row=1, column=party_idx + 2, value=parties[party_idx])

        for data_file in files_list:
            if 'numOfParties=%s.json' % str(parties[party_idx]) in data_file:
                with open(data_file, 'r') as df:
                    json_data = json.load(df, object_pairs_hook=OrderedDict)
                    for json_size_idx in range(len(json_data)):
                        for rep_idx in range(num_of_repetitions):
                            tasks_names[json_data[json_size_idx]['name']].append(
                                json_data[json_size_idx]['iteration_%s' % str(rep_idx)])
        # write data to excel
        counter = 1
        for key in tasks_names.keys():
            if len(tasks_names) > 0:
                ws.cell(row=counter + 1, column=1, value=key)
                ws.cell(row=counter + 1, column=party_idx + 2, value=sum(tasks_names[key]) / len(tasks_names[key]))
                counter += 1

        # delete all the data from the lists after finish iterate al over party data
        for val in tasks_names.values():
            val.clear()

    wb.save(results_file_name)

    send_email()

    # upload_to_git(results_file_name)


def analyze_cpu():
    files_list = glob.glob(expanduser('%s/*_cpu*.json' % results_path))
    analyze_results(files_list, 'cpu')


def analyze_comm_sent():
    files_list = glob.glob(expanduser('%s/*_commSent*.json' % results_path))
    analyze_results(files_list, 'sent')


def analyze_comm_received():
    files_list = glob.glob(expanduser('%s/*_commReceived*.json' % results_path))
    analyze_results(files_list, 'received')


def analyze_memory():
    files_list = glob.glob(expanduser('%s/*_memory*.json' % results_path))
    analyze_results(files_list, 'memory')


def analyze_all():
    analyze_cpu()
    analyze_comm_sent()
    analyze_comm_received()
    analyze_memory()
    # send_email()
    # upload_to_git()


if task_idx == '1':
    os.system('fab -f ExperimentExecute/fabfile.py collect_results:%s,%s --parallel --no-pty'
              % (remote_directory, results_path))
    # wait for all clients to download data
    time.sleep(10)
    analyze_all()
elif task_idx == '2':
    analyze_all()
