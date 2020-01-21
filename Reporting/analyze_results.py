import os
import glob
import json
import time
import smtplib
from os.path import expanduser
from os.path import basename

from openpyxl import Workbook
from openpyxl.styles import NamedStyle

from collections import OrderedDict

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


class Analyze:
    """
    The class creates reports for multi party protocols when we want to
    split the results by number of parties.
    If you want to split by different parameter, use the Elastic class
    """
    def __init__(self, protocol_config):
        """
        :type protocol_config str
        :param protocol_config: the configuration of the protocol we want to analyze
        """
        self.protocol_config = protocol_config
        self.protocol_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        self.style1 = NamedStyle(number_format='#.##')

    def download_data(self):
        """
        Download data from the remote servers
        """
        remote_directory = self.protocol_config['workingDirectory']
        results_path = self.protocol_config['resultsDirectory']
        os.system('fab -f Execution/fabfile.py collect_results:%s,%s --no-pty --parallel'
                  % (remote_directory, results_path))
        # wait for all clients to download data
        time.sleep(10)

    def send_email(self):
        """
        Sends email with the results file to the users specified in the configuration
        """
        protocol_name = self.protocol_config['protocol']
        users = list(self.protocol_config['emails'].values())
        configurations = list(self.protocol_config['configurations'].values())
        regions = list(self.protocol_config['regions.json'].values())
        address_me = 'biu.cyber.experiments@gmail.com'
        me = 'BIU Cyber Experiments <biu.cyber.experiments@gmail.com>'

        results_file_name = 'ExperimentReport/Results_%s_%s.xlsx' % (protocol_name, self.protocol_time)

        message = MIMEMultipart()
        message['Subject'] = 'Experiment results for protocol %s' % protocol_name
        message['From'] = me
        message['To'] = ', '.join(users)
        message_body = 'Results for protocol %s are attached.\n' % protocol_name
        message_body += 'The configuration(s) for this experiment are:\n\n'

        # write all the configuration to mail
        for conf in configurations:
            vals = conf.split('@')
            values_str = ''

            for val in vals:
                values_str += '%s ' % val
            message_body += '%s\n' % values_str
        # write all regions.json to mail
        message_body += 'The region(s) the experiment executed are:\n\n'
        for region in regions:
            message_body += '%s\n' % region

        message_body += '\nBIU Cyber Experiments'
        message.attach(MIMEText(message_body))

        # attach to mail all the reports file
        try:
            with open(results_file_name, 'rb') as fli:
                part = MIMEApplication(fli.read(), Name=basename(results_file_name))
                part['Content-Disposition'] = 'attachment; filename="%s"' % basename(results_file_name)
                message.attach(part)
        except EnvironmentError:
            print('Cannot attach files to mail')
            return

        try:
            with open('GlobalConfigurations/tokens.json', 'r') as tokens:
                data = json.load(tokens)
                mail_username = data['mail']['user']
                mail_password = data['mail']['password']
        except EnvironmentError:
            print('Cannot open tokens file')

        server = smtplib.SMTP('smtp.gmail.com:587')
        server.starttls()
        server.login(mail_username, mail_password)
        server.sendmail(me, users, message.as_string())
        server.quit()

    def analyze_json(self, files_list):
        """
        Analyze results files that written in JSON format
        :type files_list list
        :param files_list: list of results files location
        """
        protocol_name = self.protocol_config['protocol']
        results_path = self.protocol_config['resultsDirectory']

        parties = set()
        parties_files = OrderedDict()

        for file in files_list:
            number_of_parties = basename(file.split('*')[3])
            parties.add(int(number_of_parties))
            if number_of_parties not in parties_files.keys():
                parties_files[number_of_parties] = []
            parties_files[number_of_parties].append(file)

        parties = list(parties)  # sort parties in ascending order
        parties.sort()

        # map between number of parties and files names

        # parsing tasks names from json file
        # Assumption : all the parties measure the same tasks

        tasks_names = dict()

        # load one of the data files to receive the headers to the xlsx file
        try:
            with open(files_list[0], 'r') as f:
                data = json.load(f)

        except EnvironmentError:
            print(f'Cannot read {files_list[0]} for determine results files headers')
            return

        # init list values
        for i in range(len(data)):
            tasks_names[data[i]['name']] = list()

        num_of_repetitions = self.protocol_config['numOfInternalRepetitions']
        results_file_name = expanduser('%s/Results_%s.xlsx' % (results_path, protocol_name))

        wb = Workbook(write_only=False)  # append to file
        ws = wb.create_sheet('Results')
        wb.remove(wb['Sheet'])  # Remove the default sheet
        ws.cell(row=1, column=1, value='Phase/Number of Parties')

        files_list.sort()

        # counter = 0
        for party_idx in range(len(parties)):
            ws.cell(row=1, column=party_idx + 2, value=parties[party_idx])
            party_size = parties[party_idx]
            for data_file in parties_files[str(party_size)]:
                try:
                    with open(data_file, 'r') as df:
                        json_data = json.load(df, object_pairs_hook=OrderedDict)
                        for json_size_idx in range(len(json_data)):
                            for rep_idx in range(num_of_repetitions):
                                tasks_names[json_data[json_size_idx]['name']].append(
                                    json_data[json_size_idx]['iteration_%s' % str(rep_idx)])
                except EnvironmentError:
                    print('Cannot read results files')
                    return
            # write data to excel
            counter = 1
            for key in tasks_names.keys():
                ws.cell(row=counter + 1, column=1, value=key)
                ws.cell(row=counter + 1, column=party_idx + 2, value=sum(tasks_names[key]) / len(tasks_names[key]))
                counter += 1

            # delete all the data from the lists after finish iterate al over party data
            for val in tasks_names.values():
                val.clear()

        wb.save(results_file_name)

    def analyze_logs(self, files_list):
        """
        Analyze results files that written in log format.
        For log example file, see https://github.com/cryptobiu/MATRIX/blob/master/Assets/
        :type files_list list
        :param files_list: list of results files location
        """
        protocol_name = self.protocol_config['protocol']
        results_path = self.protocol_config['resultsDirectory']
        parties = set()
        parties_files = OrderedDict()

        # map between number of parties and file names
        for idx in range(len(files_list)):
            try:
                with open(files_list[idx]) as f:
                    data = [l.rstrip('\n') for l in f.readlines()]
                    if idx == 0:
                        # read tasks names
                        tasks_names = [l.split(':')[0] for l in data[1:]]
                        tasks_names = {t: [] for t in tasks_names}
            except EnvironmentError:
                print('Cannot read results files')

                number_of_parties = data[0]
                parties.add(number_of_parties)
                if number_of_parties not in parties_files.keys():
                    parties_files[number_of_parties] = []
                parties_files[number_of_parties].append(files_list[idx])

        parties = list(parties)
        parties.sort()  # sort parties in ascending order
        # define result file
        results_file_name = expanduser('%s/Results_%s.xlsx' % (results_path, protocol_name))

        wb = Workbook(write_only=False)  # append to file

        # write headers to excel file

        ws = wb.create_sheet('Results')
        wb.remove(wb['Sheet'])  # Remove the default sheet
        ws.cell(row=1, column=1, value='Phase/Number of Parties')
        tasks_counter = 0
        for t in tasks_names.keys():
            ws.cell(row=tasks_counter + 2, column=1, value=t)
            tasks_counter += 1

        for party_idx in range(len(parties)):
            party_size = parties[party_idx]
            ws.cell(row=1, column=party_idx + 2, value=party_size)
            for idx in range(len(parties_files[str(party_size)])):
                try:
                    with open(parties_files[str(party_size)][idx]) as data_file:
                        # get only the values for each task
                        data = [l.rstrip('\n').split(':')[1].split(',')[:-1] for l in data_file.readlines()[1:]]
                        counter = 0
                        for key in tasks_names.keys():
                            tasks_names[key].append(data[counter])
                            counter += 1
                except EnvironmentError:
                    print('Cannot read results files')

            # write the values to file for the current party_number
            data_counter = 1  # use for write the data to the correct location at the file
            for key in tasks_names.keys():
                flat_list = [item for sublist in tasks_names[key] for item in sublist]
                flat_list = list(map(int, flat_list))
                ws.cell(data_counter + 1, party_idx + 2, (sum(flat_list) / len(flat_list)))
                data_counter += 1

            for val in tasks_names.values():
                val.clear()

        # save the excel file
        wb.save(results_file_name)

    def analyze_results(self):
        """
        List all the results files and activate the relevant function
        """
        results_path = self.protocol_config['resultsDirectory']

        external_protocol = json.loads(self.protocol_config['isExternal'].lower())
        if external_protocol:
            files_list = glob.glob(expanduser('%s/*.log' % results_path))
            self.analyze_logs(files_list)
        else:
            files_list = glob.glob(expanduser('%s/*cpu*.json' % results_path))
            self.analyze_json(files_list)








